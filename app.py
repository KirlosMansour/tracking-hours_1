"""
Project Hours EAC Tracker — app.py
====================================
New in this version:
  - TEAM_MEMBERS list exposed via GET /api/team
  - Server-side state persistence:
      GET  /api/state   -> load state from data/state.json
      POST /api/state   -> save state to data/state.json
    The browser also saves to localStorage so data is kept in BOTH places.
    On load the app tries the server first; if unavailable it falls back
    to localStorage.  This means any device that can reach the server
    will see the same shared data.
"""

from flask import Flask, render_template, request, send_file, jsonify
from io import BytesIO
from datetime import datetime
import json, os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

app = Flask(__name__)

# ---------------------------------------------------------------------------
# SERVER-SIDE STATE FILE
# ---------------------------------------------------------------------------
DATA_DIR  = os.path.join(os.path.dirname(__file__), "data")
DATA_FILE = os.path.join(DATA_DIR, "state.json")

# ---------------------------------------------------------------------------
# TEAM MEMBERS
# ---------------------------------------------------------------------------
TEAM_MEMBERS = [
    "Sriram", "Garrett", "Evan", "Kirlos", "Kristi",
    "Christian", "Amanda", "Jasper", "Long"
]

# ---------------------------------------------------------------------------
# FIXED PHASE LIST
# ---------------------------------------------------------------------------
PHASES = [
    ("30% Design",  "Design Drawings"),
    ("30% Design",  "Design Studies"),
    ("60% Design",  "Design Drawings"),
    ("60% Design",  "Design Studies"),
    ("90% / IFP",   "Design Drawings"),
    ("90% / IFP",   "Design Studies"),
    ("IFC",         "Design Drawings"),
    ("IFC",         "Design Studies"),
    ("Record",      "Record Drawings"),
    ("Record",      "Design Studies"),
    ("Optional",    "BOM Generation"),
    ("Optional",    "Construction Support"),
    ("Optional",    "Commissioning Support"),
]

GROUPS = ["30% Design", "60% Design", "90% / IFP", "IFC", "Record", "Optional"]


def phases_payload():
    return [{"id": f"{i}", "group": g, "task": t} for i, (g, t) in enumerate(PHASES)]


# ---------------------------------------------------------------------------
# Shared EAC calc helpers (mirrors script.js)
# ---------------------------------------------------------------------------
def calc_row(row):
    bac  = float(row.get("bac") or 0)
    ac   = float(row.get("ac")  or 0)
    pct  = float(row.get("pct") or 0)
    etc_raw = row.get("etc", None)
    etc  = float(etc_raw) if etc_raw not in (None, "") else None

    ev   = bac * (pct / 100.0)
    cpi  = (ev / ac) if ac > 0 else None

    eac1 = (bac / cpi) if cpi else None
    eac2 = ac + (bac - ev)
    eac3 = (ac + etc) if etc is not None else None

    if etc is not None:
        eac_main, method = eac3, "Method 3"
    elif eac1 is not None:
        eac_main, method = eac1, "Method 1"
    else:
        eac_main, method = eac2, "Method 2"

    return {
        "ev": ev, "cpi": cpi, "eac1": eac1, "eac2": eac2, "eac3": eac3,
        "eac_main": eac_main, "method": method, "variance": eac_main - bac,
    }


def status_for(variance, bac):
    if variance <= 0:                              return "On Track"
    if bac > 0 and variance <= 0.10 * bac:        return "Watch"
    return "Over Budget"


def project_totals(project):
    rows = project.get("phases", [])
    total_bac = sum(float(r.get("bac") or 0) for r in rows)
    total_ac  = sum(float(r.get("ac")  or 0) for r in rows)
    total_ev  = total_eac = 0.0
    for r in rows:
        c = calc_row(r)
        total_ev  += c["ev"]
        total_eac += c["eac_main"]
    overall_pct = (total_ev / total_bac * 100.0) if total_bac > 0 else 0.0
    overall_cpi = (total_ev / total_ac)           if total_ac  > 0 else None
    total_variance = total_eac - total_bac
    return {
        "total_bac": total_bac, "total_ac": total_ac, "overall_pct": overall_pct,
        "total_ev": total_ev, "overall_cpi": overall_cpi, "total_eac": total_eac,
        "total_variance": total_variance, "status": status_for(total_variance, total_bac),
    }


# ---------------------------------------------------------------------------
# Routes — meta
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/phases")
def api_phases():
    return jsonify(phases_payload())

@app.route("/api/groups")
def api_groups():
    return jsonify(GROUPS)

@app.route("/api/team")
def api_team():
    return jsonify(TEAM_MEMBERS)


# ---------------------------------------------------------------------------
# Routes — server-side state persistence
# ---------------------------------------------------------------------------
@app.route("/api/state", methods=["GET"])
def get_state():
    """Return the saved state from disk, or null if nothing saved yet."""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return jsonify(json.load(f))
        except Exception as e:
            print(f"[state] read error: {e}")
    return jsonify(None)


@app.route("/api/state", methods=["POST"])
def save_state():
    """Persist the full state JSON sent by the browser to disk."""
    try:
        data = request.get_json(force=True)
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return jsonify({"ok": True})
    except Exception as e:
        print(f"[state] write error: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------------------------------------------------------------------------
# Routes — Excel export
# ---------------------------------------------------------------------------
@app.route("/api/export/excel", methods=["POST"])
def export_excel():
    data     = request.get_json(force=True)
    projects = data.get("projects", [])

    wb = Workbook()
    wb.remove(wb.active)

    hfill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    hfont  = Font(color="FFFFFF", bold=True)
    gfill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Dashboard sheet
    sws = wb.create_sheet("Dashboard")
    sh  = ["Project Name", "Total Planned Hours", "Total Actual Hours",
           "Percent Complete", "Total EAC", "Variance", "Status"]
    sws.append(sh)
    for c in range(1, len(sh) + 1):
        cell = sws.cell(row=1, column=c)
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center")
    for project in projects:
        t = project_totals(project)
        sws.append([project.get("name", ""), round(t["total_bac"], 1), round(t["total_ac"], 1),
                    f'{t["overall_pct"]:.1f}%', round(t["total_eac"], 1),
                    round(t["total_variance"], 1), t["status"]])
    for col in range(1, len(sh) + 1):
        sws.column_dimensions[get_column_letter(col)].width = 20

    # Time log sheet (if any entries exist)
    all_logs = data.get("timeLogs", [])
    if all_logs:
        lws = wb.create_sheet("Time Log")
        lh  = ["Date", "Member", "Hours", "Project", "Stage", "Type", "Logged At"]
        lws.append(lh)
        for c in range(1, len(lh) + 1):
            cell = lws.cell(row=1, column=c)
            cell.fill, cell.font = hfill, hfont
            cell.alignment = Alignment(horizontal="center")
        for entry in all_logs:
            lws.append([entry.get("date",""), entry.get("member",""),
                        entry.get("hours",0), entry.get("projectName",""),
                        entry.get("stage",""), entry.get("type",""),
                        entry.get("addedAt","")])
        for col in range(1, len(lh) + 1):
            lws.column_dimensions[get_column_letter(col)].width = 20

    # Team Allocations sheet
    tws = wb.create_sheet("Team Allocations")
    th  = ["Project", "Status", "Member", "Allocated Hours", "Share %",
           "Full Hrs/Day", "Workload %", "Effective Hrs/Day", "Working Days"]
    tws.append(th)
    for c in range(1, len(th) + 1):
        cell = tws.cell(row=1, column=c)
        cell.fill, cell.font = hfill, hfont
        cell.alignment = Alignment(horizontal="center")
    for project in projects:
        allocs     = project.get("teamAllocations", [])
        total_hrs  = sum(float(a.get("hours", 0) or 0) for a in allocs)
        start, end = project.get("startDate",""), project.get("endDate","")
        w_days     = None
        if start and end:
            from datetime import date as dt, timedelta
            s, e = dt.fromisoformat(start), dt.fromisoformat(end)
            w_days = sum(1 for d in (s + timedelta(n) for n in range((e-s).days+1)) if d.weekday() < 5)
        status = "Closed" if project.get("isClosed") else "Active"
        for a in allocs:
            hrs    = float(a.get("hours", 0) or 0)
            wl_pct = float(a.get("workloadPct", 100) or 100)
            share  = round(hrs / total_hrs * 100, 1) if total_hrs > 0 else 0
            hpd    = round(hrs / w_days, 2) if w_days else ""
            eff    = round(hpd * wl_pct / 100, 2) if w_days else ""
            tws.append([project.get("name",""), status, a.get("member",""),
                        hrs, f"{share}%", hpd, f"{wl_pct}%", eff,
                        w_days or ""])
    for col in range(1, len(th) + 1):
        tws.column_dimensions[get_column_letter(col)].width = 20

    # One sheet per project
    cols = ["Group", "Task", "Planned (BAC)", "Actual (AC)", "% Complete",
            "ETC", "Earned (EV)", "CPI", "EAC Method 1", "EAC Method 2",
            "EAC Method 3", "EAC (used)", "Variance"]
    for project in projects:
        ws = wb.create_sheet((project.get("name") or "Project")[:28])
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill, cell.font = hfill, hfont
            cell.alignment = Alignment(horizontal="center")
        for r in project.get("phases", []):
            c = calc_row(r)
            ws.append([r.get("group",""), r.get("task",""),
                       r.get("bac") or 0, r.get("ac") or 0, r.get("pct") or 0,
                       r.get("etc") if r.get("etc") not in (None,"") else "",
                       round(c["ev"],1),
                       round(c["cpi"],2) if c["cpi"] is not None else "N/A",
                       round(c["eac1"],1) if c["eac1"] is not None else "N/A",
                       round(c["eac2"],1),
                       round(c["eac3"],1) if c["eac3"] is not None else "",
                       round(c["eac_main"],1), round(c["variance"],1)])
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(cols)):
            for cell in row:
                cell.border = border
        t = project_totals(project)
        ws.append([])
        ws.append(["TOTAL","",round(t["total_bac"],1),round(t["total_ac"],1),
                   f'{t["overall_pct"]:.1f}%',"",round(t["total_ev"],1),
                   round(t["overall_cpi"],2) if t["overall_cpi"] else "N/A",
                   "","","",round(t["total_eac"],1),round(t["total_variance"],1)])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = gfill
        for col in range(1, len(cols)+1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        if project.get("notes"):
            ws.append([])
            ws.append(["Notes:", project["notes"]])

    out = BytesIO()
    wb.save(out); out.seek(0)
    fname = f"project_hours_eac_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Routes — PDF export
# ---------------------------------------------------------------------------
@app.route("/api/export/pdf", methods=["POST"])
def export_pdf():
    data     = request.get_json(force=True)
    projects = data.get("projects", [])

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.4*inch, rightMargin=0.4*inch,
                            topMargin=0.4*inch,  bottomMargin=0.4*inch)
    styles = getSampleStyleSheet()
    ts  = ParagraphStyle("T",  parent=styles["Title"],   fontSize=18)
    h2  = ParagraphStyle("H2", parent=styles["Heading2"], spaceBefore=10)
    nor = styles["Normal"]

    elements = []
    elements.append(Paragraph("Project Hours EAC Tracker — Report", ts))
    elements.append(Paragraph(datetime.now().strftime("Generated %B %d, %Y %H:%M"), nor))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("Dashboard Summary", h2))
    dd   = [["Project","Planned Hrs","Actual Hrs","% Complete","Total EAC","Variance","Status"]]
    scol = {"On Track": colors.HexColor("#16a34a"),
            "Watch":    colors.HexColor("#ca8a04"),
            "Over Budget": colors.HexColor("#dc2626")}
    rstatus = []
    for p in projects:
        t = project_totals(p)
        dd.append([p.get("name",""), f'{t["total_bac"]:.0f}', f'{t["total_ac"]:.0f}',
                   f'{t["overall_pct"]:.1f}%', f'{t["total_eac"]:.0f}',
                   f'{t["total_variance"]:+.0f}', t["status"]])
        rstatus.append(t["status"])
    dtbl = Table(dd, repeatRows=1)
    scmds = [("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),
             ("TEXTCOLOR",(0,0),(-1,0),colors.white),
             ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
             ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#cccccc")),
             ("FONTSIZE",(0,0),(-1,-1),9),("ALIGN",(1,0),(-1,-1),"CENTER")]
    for i, st in enumerate(rstatus, 1):
        scmds += [("TEXTCOLOR",(6,i),(6,i),scol.get(st,colors.black)),
                  ("FONTNAME",(6,i),(6,i),"Helvetica-Bold")]
    dtbl.setStyle(TableStyle(scmds))
    elements += [dtbl, PageBreak()]

    pcols = ["Group","Task","BAC","AC","%Comp","ETC","EV","CPI",
             "EAC-1","EAC-2","EAC-3","EAC(used)","Var"]
    for p in projects:
        elements.append(Paragraph(p.get("name","Project"), h2))
        t = project_totals(p)
        elements.append(Paragraph(
            f'Overall % Complete: {t["overall_pct"]:.1f}% | '
            f'CPI: {(f"{t["overall_cpi"]:.2f}" if t["overall_cpi"] else "N/A")} | '
            f'Total EAC: {t["total_eac"]:.0f} hrs | '
            f'Variance: {t["total_variance"]:+.0f} hrs | Status: {t["status"]}', nor))
        elements.append(Spacer(1, 6))
        td = [pcols]
        for r in p.get("phases", []):
            c = calc_row(r)
            td.append([r.get("group",""), r.get("task",""),
                       f'{(r.get("bac") or 0):.0f}', f'{(r.get("ac") or 0):.0f}',
                       f'{(r.get("pct") or 0):.0f}%',
                       f'{r.get("etc"):.0f}' if r.get("etc") not in (None,"") else "-",
                       f'{c["ev"]:.0f}',
                       f'{c["cpi"]:.2f}' if c["cpi"] else "N/A",
                       f'{c["eac1"]:.0f}' if c["eac1"] else "N/A",
                       f'{c["eac2"]:.0f}',
                       f'{c["eac3"]:.0f}' if c["eac3"] else "-",
                       f'{c["eac_main"]:.0f}', f'{c["variance"]:+.0f}'])
        ptbl = Table(td, repeatRows=1)
        ptbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#dddddd")),
            ("FONTSIZE",(0,0),(-1,-1),7.5),("ALIGN",(2,0),(-1,-1),"CENTER"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f5f7fa")]),
        ]))
        elements.append(ptbl)
        if p.get("notes"):
            elements += [Spacer(1,6), Paragraph(f'<b>Notes:</b> {p["notes"]}', nor)]

        # Team Allocations for this project
        allocs    = p.get("teamAllocations", [])
        total_hrs = sum(float(a.get("hours",0) or 0) for a in allocs)
        start, end = p.get("startDate",""), p.get("endDate","")
        w_days = None
        if start and end:
            from datetime import date as dt2, timedelta
            s2, e2 = dt2.fromisoformat(start), dt2.fromisoformat(end)
            w_days = sum(1 for dd in (s2+timedelta(n) for n in range((e2-s2).days+1)) if dd.weekday()<5)
        if allocs:
            elements.append(Spacer(1, 8))
            elements.append(Paragraph("Team Allocations", h2))
            closed_txt = " <b>[CLOSED]</b>" if p.get("isClosed") else ""
            elements.append(Paragraph(f'Status: {"Closed" if p.get("isClosed") else "Active"}{closed_txt}', nor))
            atd = [["Member","Alloc. Hrs","Share %","Full Hrs/Day","Workload %","Eff. Hrs/Day"]]
            for a in allocs:
                hrs   = float(a.get("hours",0) or 0)
                wlp   = float(a.get("workloadPct",100) or 100)
                share = f'{hrs/total_hrs*100:.1f}%' if total_hrs>0 else "0%"
                hpd   = f'{hrs/w_days:.2f}' if w_days else "—"
                eff   = f'{hrs/w_days*wlp/100:.2f}' if w_days else "—"
                atd.append([a.get("member",""), f'{hrs:.0f}', share, hpd, f'{wlp:.0f}%', eff])
            atbl = Table(atd, repeatRows=1)
            atbl.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E78")),
                ("TEXTCOLOR",(0,0),(-1,0),colors.white),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#dddddd")),
                ("FONTSIZE",(0,0),(-1,-1),8.5),
                ("ALIGN",(1,0),(-1,-1),"CENTER"),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f5f7fa")]),
            ]))
            elements.append(atbl)

        elements.append(PageBreak())

    if elements and isinstance(elements[-1], PageBreak):
        elements.pop()
    doc.build(elements)
    buf.seek(0)
    fname = f"project_hours_eac_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/pdf")


# ---------------------------------------------------------------------------
# Routes — Time Log PDF export
# ---------------------------------------------------------------------------
@app.route("/api/export/pdf/timelog", methods=["POST"])
def export_pdf_timelog():
    """Export time log entries as a standalone PDF, separate from project data."""
    data      = request.get_json(force=True)
    time_logs = data.get("timeLogs", [])

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.4*inch, rightMargin=0.4*inch,
                            topMargin=0.4*inch,  bottomMargin=0.4*inch)
    styles = getSampleStyleSheet()
    ts  = ParagraphStyle("T",  parent=styles["Title"],   fontSize=18)
    h2  = ParagraphStyle("H2", parent=styles["Heading2"], spaceBefore=8)
    nor = styles["Normal"]

    elements = []
    elements.append(Paragraph("Time Log Report", ts))
    elements.append(Paragraph(datetime.now().strftime("Generated %B %d, %Y %H:%M"), nor))
    elements.append(Spacer(1, 10))

    # Summary by member
    if time_logs:
        member_totals = {}
        for e in time_logs:
            m = e.get("member", "Unknown")
            member_totals[m] = member_totals.get(m, 0) + float(e.get("hours", 0))

        elements.append(Paragraph("Summary by Team Member", h2))
        sum_data = [["Member", "Total Hours"]]
        for m, h in sorted(member_totals.items()):
            sum_data.append([m, f"{h:.1f}"])
        sum_data.append(["TOTAL", f"{sum(member_totals.values()):.1f}"])

        stbl = Table(sum_data, colWidths=[2.5*inch, 1.5*inch])
        stbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0),  (-1,0),  colors.HexColor("#1F4E78")),
            ("TEXTCOLOR",  (0,0),  (-1,0),  colors.white),
            ("FONTNAME",   (0,0),  (-1,0),  "Helvetica-Bold"),
            ("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"),
            ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#D9E1F2")),
            ("GRID",       (0,0),  (-1,-1), 0.5, colors.HexColor("#cccccc")),
            ("FONTSIZE",   (0,0),  (-1,-1), 9),
            ("ALIGN",      (1,0),  (1,-1),  "CENTER"),
        ]))
        elements.append(stbl)
        elements.append(Spacer(1, 14))

    # Full log table
    elements.append(Paragraph("Full Hours Log", h2))
    if not time_logs:
        elements.append(Paragraph("No time log entries recorded yet.", nor))
    else:
        cols = ["Date", "Member", "Hours", "Project", "Stage", "Type", "Logged At"]
        tdata = [cols]
        for e in time_logs:
            added = e.get("addedAt", "")[:16].replace("T", " ") if e.get("addedAt") else ""
            tdata.append([
                e.get("date", ""),
                e.get("member", ""),
                f'{float(e.get("hours", 0)):.1f}',
                e.get("projectName", ""),
                e.get("stage", ""),
                e.get("type", ""),
                added,
            ])
        ltbl = Table(tdata, repeatRows=1)
        ltbl.setStyle(TableStyle([
            ("BACKGROUND",     (0,0), (-1,0),  colors.HexColor("#1F4E78")),
            ("TEXTCOLOR",      (0,0), (-1,0),  colors.white),
            ("FONTNAME",       (0,0), (-1,0),  "Helvetica-Bold"),
            ("GRID",           (0,0), (-1,-1), 0.4, colors.HexColor("#dddddd")),
            ("FONTSIZE",       (0,0), (-1,-1), 8.5),
            ("ALIGN",          (2,0), (2,-1),  "CENTER"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [colors.white, colors.HexColor("#f5f7fa")]),
        ]))
        elements.append(ltbl)

    doc.build(elements)
    buf.seek(0)
    fname = f"time_log_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/pdf")


if __name__ == "__main__":
    app.run(debug=True, port=5000)
